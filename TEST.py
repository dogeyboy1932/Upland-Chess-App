import http.client
import json
from Upland.get_user_balance import UpdateBalance
# from openpyxl import load_workbook
# from Chess.FIXED_CHESS_VARIABLES import cfilepath

# def AppendInitial():
#     workbook = load_workbook(cfilepath)
#     worksheet = workbook['Sheet']

#     data = ["gameID", "challenger", "rating", "wager", "link", "escrowID", "accepter"]

#     worksheet.append(data)

#     workbook.save(cfilepath)
#     workbook.close()




def GetEscrowContainer(eid):
    conn = http.client.HTTPSConnection("api.sandbox.upland.me")
    payload = json.dumps({
        "expirationPeriodHours": 24
    })

    headers = {
        'Authorization': 'Basic MjMyOmFkMzMxMDkxLTQ3NjItNGZlMS1iNDBmLTFkNGNhMGQwMmQ5Zg==',
        'Cookie': 'sticky-session-1=1701556690.435.2069.742375|9a5cc3e4d08faea009d8e16f5c97bee9'
    }

    url = "/containers/" + str(eid)

    conn.request("GET", url, payload, headers)
    res = conn.getresponse()
    data = json.loads(res.read().decode("utf-8"))

    # print(data)

    return data


# def JoinEscrow(bearerToken, containerId, upxAmount):
#     conn = http.client.HTTPSConnection("api.sandbox.upland.me")
#     payload = json.dumps({
#         "containerId": containerId,
#         "upxAmount": upxAmount,
#         "sparkAmount": 0,
#         "assets": [],
#     })

#     bearer = 'Bearer ' + str(bearerToken)

#     headers = {
#         # 'Authorization': str(bearer),
#         'Authorization': bearer,
#         'Content-Type': 'application/json',
#         'Cookie': 'sticky-session-1=1699553168.246.2069.619172|aebf5e9dc298523c710b3cfe411c6704'
#     }

#     conn.request("POST", "/developers-api/User/join", payload, headers)
#     res = conn.getresponse()

#     if res.status == 200 or res.status == 201:
#         data = json.loads(res.read().decode("utf-8"))
#         print("Joined Escrow! Transaction Hash:", data["transactionId"])
#         return "success"
#     else:
#         print(f'Request failed with status code {res.status}')
#         return "error"


# def run():
#     eid = 3968
#     # eid = CreateEscrowContainer()
#     # print(eid)

#     # bearer = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiYTUwZWIxZGEtMjI1ZS00MWY5LWFhM2ItN2M5NzQwZjUwZmY1IiwiaWF0IjoxNzA0OTk0MDk2fQ.xxnNviAMmuzmPz2R9cniCJ2BwZOTM6ya823dOjqhhAw'

#     # JoinEscrow(bearer, eid, 50)
#     print(GetEscrowContainer(eid))

# run()

# import http.client
# import json

# conn = http.client.HTTPSConnection("api.sandbox.upland.me")
# payload = json.dumps({
#   "containerId": str(3191),
#   "upxAmount": 10,
#   "sparkAmount": 0,
#   "assets": []
# })
# headers = {
#   'Content-Type': 'application/json',
#   'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiYzdjZGYxZmEtNTIxMi00MTI1LTg5ZDctNThiNDk1OTVkZDQzIiwiaWF0IjoxNzA0OTE4ODI5fQ.UdH-iKZddS4pKnfYBuKK0AfYGBfP-cEXlaplsMPnn9o',
#   'Cookie': 'sticky-session-1=1704817779.488.2071.595873|9a5cc3e4d08faea009d8e16f5c97bee9'
# }
# conn.request("POST", "/developers-api/User/join", payload, headers)
# res = conn.getresponse()
# data = res.read()
# # print(res.headers)
# print(data.decode("utf-8"))
# print("HERE")

# from Upland.FIXED_VARIABLES import conn
# import json

# def JoinEscrow(bearerToken, containerId, upxAmount):
#     payload = json.dumps({
#         "containerId": containerId,
#         "upxAmount": upxAmount,
#         "sparkAmount": 0,
#         "assets": [],
#     })

#     bearer = 'Bearer ' + str(bearerToken)

#     headers = {
#         'Authorization': str(bearer),
#         # 'Authorization': bearer,
#         # 'Content-Type': 'application/json',
#         'Cookie': 'sticky-session-1=1699553168.246.2069.619172|aebf5e9dc298523c710b3cfe411c6704'
#     }

#     conn.request("POST", "/developers-api/User/join", payload, headers)
#     res = conn.getresponse()

#     if res.status == 200 or res.status == 201:
#         data = json.loads(res.read().decode("utf-8"))
#         print("Joined Escrow! Transaction Hash:", data["transactionId"])
#         return "success"
#     else:
#         print(f'Request failed with status code {res.status}')
#         print(res.msg)
#         return "error"


# def run():
#     eid = 3191
#     # eid = CreateEscrowContainer()
#     print(eid)

#     bearer = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiYTcyNzE2NDEtZTcxMS00YjBiLWE5NWYtMjEwNTA3Y2JlZGQyIiwiaWF0IjoxNzA0MzA1OTU5fQ.suGtQmS0pEEyxgJmqDXkd9ZKf0P4y5_PrTIs9y6_vJA'

#     JoinEscrow(bearer, eid, 100)

# run()

# # l3cn1ph243wu

from Upland.FIXED_VARIABLES import conn
from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook
import json

def UpdateBalance(upland_access_token):
    payload = ""
    headers = {
        'Authorization': f'Bearer {upland_access_token}',
    }

    while(True):
        try:
            conn.request("GET", "/user/balances", payload, headers)
            break
        except:
            print("NOT WORKING")

    res = conn.getresponse()
    data = json.loads(res.read().decode("utf-8"))

    print(data)

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == upland_access_token:
            worksheet[i][3].value = data['availableUpx']

def GetUserBalance(upland_access_token):
    payload = ""
    headers = {
        'Authorization': f'Bearer {upland_access_token}',
    }

    conn.request("GET", "/user/balances", payload, headers)
    res = conn.getresponse()
    data = json.loads(res.read().decode("utf-8"))

    if res.status == 200:
        return data['availableUpx']
    else:
        print(f'Request failed with status code {res.status}')

def run():
    upland_access_token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiIwOTQ4ZGE1MC04N2Q0LTExZWUtYjBjMi02MzM4M2I3OTUzNjAiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiM2MzNWQ4YTMtYTAwOS00MmM2LWJmN2UtM2RiN2E5MTlmYmI2IiwiaWF0IjoxNzEyMzQ2MjkyfQ.nDWgq4Co0JMXht5O6U0OUY-NhwmQn1OkGCffJGE6HBg"

    print(UpdateBalance(upland_access_token))

    # print(GetUserBalanceOnSheet("dogeyboy19"))


# run()


import http.client

conn = http.client.HTTPSConnection("api.sandbox.upland.me")

payload = ''
headers = {
  'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiIwOTQ4ZGE1MC04N2Q0LTExZWUtYjBjMi02MzM4M2I3OTUzNjAiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiNWRmNDQ1MTUtZmMzNy00Mjk1LWI4MTktZjgzOGIwNThiZjFlIiwiaWF0IjoxNzEyMzQ2NDcyfQ.3yhchzHnGPLA1U0d8bVw528tcBuDBksNTwP3ZiOB7ic',
  'Cookie': 'sticky-session-1=1712345152.12.32.298740|9a5cc3e4d08faea009d8e16f5c97bee9'
}
conn.request("GET", "/developers-api/user/balances", payload, headers)
res = conn.getresponse()
data = res.read()
print(data.decode("utf-8"))




# conn = http.client.HTTPSConnection("api.sandbox.upland.me/developers-api")

# payload = ''
# headers = {
#   'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiODVmZTQwMTUtYTYwZi00MWQ2LTk0ZmMtMmFlOGIzNWUyMWJlIiwiaWF0IjoxNzEyMzQyOTI5fQ.ka92uTcZJ4XMFYwVNbOqFMcHG3ppPZ7G6IHH9iaqR7U',
#   'Cookie': 'sticky-session-1=1712345152.12.32.298740|9a5cc3e4d08faea009d8e16f5c97bee9'
# }
# conn.request("GET", "user/balances", payload, headers)
# res = conn.getresponse()
# data = res.read()
# print(data.decode("utf-8"))