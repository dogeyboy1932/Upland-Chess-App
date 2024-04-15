from FIXED_VARIABLES import cfilepath
from openpyxl import load_workbook
from Upland.refund_escrow import RefundEscrowContainer
from Upland.get_escrow_container import GetEscrowContainer

def ChallengeCanceled(link):  # <- Accept Button clicked
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']
    
    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == link:
            worksheet[i][6].value = False

    workbook.save(cfilepath)
    workbook.close()