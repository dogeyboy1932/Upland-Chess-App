from openpyxl import load_workbook
import pandas as pd
from Chess.FIXED_CHESS_VARIABLES import cfilepath
from Chess.isGameFinished import isGameFinished
from Chess.game_ended import gameEnded

workbook = load_workbook(cfilepath)
worksheet = workbook['Sheet']


def ChallengeAccepted(link, challenger):  # <- Accept Button clicked
    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == link:
            worksheet[i][6].value = True

    workbook.save(cfilepath)
    workbook.close()

    ############################
    # INSTRUCTION
    # AFTER BOTH PLAYERS ACCEPT GAME WILL START...WHEN GAME STARTS DO THIS NEXT
    # Game needs 2 players to start it
    # Need function to let us know when the game starts!!
    # Get lichessId of the challenge creator (Player 1) [Query spreadsheet w/ gameID]
    # challenger cannot be the same lichessID as Player 1 -> otherwise throw a rejection pop-up when clicked
    # CHALLENGER MUST HAVE A VALID LICHESS ID
    # Add challenger lichessID to the challenge [params gameID & challenger]
    # Need to run stream_board_game_state once it does start to get the details of players
    # Call QueryForUplandID and extract respective uplandID of both lichess accounts
    #
    #```
    # Challenger will also be made to join the escrow account when he clicks accept button
    #
    # uplandID = QueryForUplandID(challenger)
    # bearer = GetBearerToken(uplandID)
    #
    # challengeIdx = GetChallengeIdx(gameID)
    # eid = chessWorksheet[challengeIdx][5]
    #
    # wager = chessWorksheet[challengeIdx][3]
    #
    # joinEscrow(bearer, eid, wager)
    # ```
    #
    # POP-UP THAT DISPLAYS (GO TO UPLAND ACCOUNT TO ACCEPT WAGER TRANSACTION)
    # TAKE ME TO CHESS GAME BUTTON -> When clicked get redirect to chessgame link (chessWorksheet[challengeIdx][4])
    # User must hit reset button for updates
    # balances are deducted from accounts

    # Need function to let us know when the game ends!! This triggers gameEnded function
    ############################

# EID 2241 GameID pf9XoRNn