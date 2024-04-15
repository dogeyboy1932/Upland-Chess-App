import pandas as pd
from openpyxl import load_workbook

from FIXED_VARIABLES import filepath, cfilepath
from Upland.SpreadsheetEditing.query_spreadsheet import QueryForUplandID
from Upland.SpreadsheetEditing.get_user_balance import UpdateBalance

# FRONTEND DEPENDENT <- Translates spreadsheet into frontend database

def Iterate():
    df = pd.read_excel(cfilepath, usecols='B:H,')
    res = []

    for index, row in df.iterrows():
        uplandID = QueryForUplandID(row.iloc[0])
        res.append([row.iloc[0], row.iloc[1], row.iloc[2], row.iloc[3], uplandID, row.iloc[5], row.iloc[6]])
    
    return res


def UpdateBalances():
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(2, worksheet.max_row + 1):
        if (worksheet[i][4].value != None):
            # print(worksheet[i][4].value)
            UpdateBalance(worksheet[i][4].value)

def run():
    Iterate()


# run()
