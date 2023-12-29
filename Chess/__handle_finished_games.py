from openpyxl import load_workbook
import pandas as pd
from Chess.FIXED_CHESS_VARIABLES import cfilepath
from Chess.isGameFinished import isGameFinished
from Chess.game_ended import gameEnded

workbook = load_workbook(cfilepath)
worksheet = workbook['Sheet']

def FindAndRemoveRow(gameID):
    # workbook = load_workbook(cfilepath)
    # worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            worksheet.delete_rows(i)
            break

    workbook.save(cfilepath)
    workbook.close()


def HandleFinishedGames():
    # workbook = load_workbook(cfilepath)
    # worksheet = workbook['Sheet']

    finishedGames = []

    for i in range(1, worksheet.max_row + 1):
        if isGameFinished(worksheet[i][0].value):
            finishedGames.append(worksheet[i][0].value)

    if not finishedGames:
        return -1

    for game in finishedGames:
        gameEnded(game)
        FindAndRemoveRow(game)
    

    # isGameValid() if not, remove from list and refund escrow container

    # for i in range(1, worksheet.max_row + 1):
    #     if not isGameValid(worksheet[i][0].value):
    #         refundEscrowContainer(worksheet[i][5].value)
    #         FindAndRemoveRow(worksheet[i][0].value)

    
    return finishedGames


def run():
    df = pd.read_excel(cfilepath)
    print(df)

    HandleFinishedGames()


# run()




# def RemovedFinishedGames(finishedGames):
# print("GameID", gameID)
# print("comparable", worksheet[i][0].value)

# print(i)
# print(worksheet[i][0].value)

# print(game)

# import pandas as pd
# df = pd.read_excel(filepath)
# print(df)
    
# df = pd.read_excel(cfilepath)
# print(df)
    
# print("1 ",worksheet[i][0].value)
# print("2 ", gameID)

# workbook = load_workbook(cfilepath)
# worksheet = workbook['Sheet']

# workbook.save(cfilepath)
# workbook.close()
