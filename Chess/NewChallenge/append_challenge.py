from openpyxl import load_workbook

from FIXED_VARIABLES import cfilepath

from Chess.get_lichess_info import GetLichessRating

from Upland.Escrow.create_escrow_container import CreateEscrowContainer


def AppendChallenge(challenger, wager, thisGame):
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    # Grabbing Details
    gameID = thisGame['challenge']['id']
    rating = GetLichessRating(challenger, "rapid")
    link = thisGame['challenge']['url']
    escrowID = CreateEscrowContainer()

    # Making data
    data = [gameID, challenger, rating, wager, link, escrowID, "NO", "blank", "NO"]

    # Appending data to spreadsheet
    worksheet.append(data)

    workbook.save(cfilepath)
    workbook.close()

    return gameID


def AppendChallengeHeader():
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    data = ["gameID", "challenger", "rating", "wager", "link", "escrowID", "accepted?", "accepter", "readyStatus"]

    worksheet.append(data)

    workbook.save(cfilepath)
    workbook.close()