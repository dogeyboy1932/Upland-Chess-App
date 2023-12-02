from Chess.FIXED_CHESS_VARIABLES import cfilepath
from openpyxl import load_workbook

def GetAllChallenges():
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    challenges = []

    for i in range(1, worksheet.max_row + 1):
        challenges.append(worksheet[i])

    return challenges
