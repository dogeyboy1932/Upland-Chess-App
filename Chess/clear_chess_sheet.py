from openpyxl import load_workbook
from FIXED_CHESS_VARIABLES import cfilepath


def run():
    workbook = load_workbook(cfilepath)
    sheet = workbook['Sheet']

    print("Maximum rows before removing:", sheet.max_row)

    if sheet.max_row != 0:
        sheet.delete_rows(2, sheet.max_row)

    print("Maximum rows after removing:", sheet.max_row)

    workbook.save(cfilepath)
    workbook.close()


run()
