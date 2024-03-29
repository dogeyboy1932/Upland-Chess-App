from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook
from Chess.game_winner import GameWinner

# import pandas as pd

def QueryForUplandID(lichessID):
    if lichessID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == lichessID:
            return worksheet[i][1].value

    return -1


def run():
    gameID = "RLyFq9MX"
    winner = GameWinner(gameID=gameID)
    corr_uplandID = QueryForUplandID(winner)

    print(corr_uplandID)


# run()


# df = pd.read_excel(filepath)
# print(df)
    
# print(worksheet[i][0].value)
# print(lichessID)
