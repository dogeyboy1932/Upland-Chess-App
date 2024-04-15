from openpyxl import load_workbook
from FIXED_VARIABLES import cfilepath
from Chess.get_chess_info import GetLichessRating
from Upland.create_escrow_container import CreateEscrowContainer

import pandas as pd

def AppendChallenge(challenger, wager, thisGame):
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    # Grabbing Details
    gameID = thisGame['challenge']['id']
    rating = GetLichessRating(challenger, "rapid")
    link = thisGame['challenge']['url']
    escrowID = CreateEscrowContainer()

    # Making data
    data = [gameID, challenger, rating, wager, link, escrowID, False, "blank"]

    # Appending data to spreadsheet
    worksheet.append(data)

    workbook.save(cfilepath)
    workbook.close()

    return gameID

def AppendInitial():
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    data = ["gameID", "challenger", "rating", "wager", "link", "escrowID", "accepted?", "accepter"]

    worksheet.append(data)

    workbook.save(cfilepath)
    workbook.close()