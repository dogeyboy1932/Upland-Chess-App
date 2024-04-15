from openpyxl import load_workbook
import pandas as pd
from FIXED_VARIABLES import cfilepath
from Chess.isGameFinished import isGameFinished
from Chess.game_ended import gameEnded
from Upland.refund_escrow import RefundEscrowContainer


def ChallengeDeleted(link):  # <- Delete Button clicked
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']
    
    challengeIdx = -1
    
    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == link:
            challengeIdx = i
            break
    
    eid = worksheet[challengeIdx][5].value
    RefundEscrowContainer(eid)

    worksheet.delete_rows(challengeIdx)

    workbook.save(cfilepath)
    workbook.close()


def FindAndRemoveRow(gameID):
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            worksheet.delete_rows(i)
            break

    workbook.save(cfilepath)
    workbook.close()
            

def HandleFinishedGames():
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    finishedGames = []

    for i in range(2, worksheet.max_row + 1):
        if isGameFinished(worksheet[i][0].value):
            finishedGames.append(worksheet[i][0].value)

    if not finishedGames:
        return -1
    
    for game in finishedGames:
        if (gameEnded(game) == -1):
            continue

        FindAndRemoveRow(game)

    return finishedGames


def run():
    df = pd.read_excel(cfilepath)
    print(df)

    HandleFinishedGames()


# run()
