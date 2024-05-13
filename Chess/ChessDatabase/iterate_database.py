import pandas as pd
from openpyxl import load_workbook

from FIXED_VARIABLES import filepath, cfilepath

from Upland.get_user_balance import GetUserBalance
from Upland.SpreadsheetEditing.query_spreadsheet import QueryForUplandID


def Iterate():  # <- Translates spreadsheet into frontend database
    df = pd.read_excel(cfilepath, usecols='B:I,')
    table = []

    for index, row in df.iterrows():
        uplandID = QueryForUplandID(row.iloc[0])
        table.append([row.iloc[0], row.iloc[1], row.iloc[2], row.iloc[3], uplandID, str(row.iloc[5]), row.iloc[6], row.iloc[7], row.iloc[4]])
    
    return table


def UpdateBalance(upland_access_token):
    availableUpx = GetUserBalance(upland_access_token)

    if availableUpx == -1: return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == upland_access_token:
            worksheet[i][3].value = availableUpx
            break

    workbook.save(filepath)
    workbook.close()


def UpdateBalances():
    worksheet = load_workbook(filepath)['Sheet']

    for i in range(2, worksheet.max_row + 1):
        UpdateBalance(worksheet[i][4].value)
            