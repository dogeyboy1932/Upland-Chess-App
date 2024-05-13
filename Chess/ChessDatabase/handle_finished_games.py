import pandas as pd
from openpyxl import load_workbook

from FIXED_VARIABLES import cfilepath

from Chess.ChessDatabase.game_ended import gameEnded, isGameFinished

from Upland.Escrow.get_escrow_container import GetEscrowContainer
from Upland.SpreadsheetEditing.query_spreadsheet import GetChallengeIdx


def FindAndRemoveRow(gameID):  # THIS REMOVES THE ROW IN THE CHALLENGES DATABASE CONTAINING THE GAMEID
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            worksheet.delete_rows(i)
            break

    workbook.save(cfilepath)
    workbook.close()


def MarkResolving(gameID): # THIS INDICATES THAT THE CHESS sGAME IS OVER AND THE ESCROW IS BEING RESOLVED
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            worksheet[i][6].value = "COMPLETED"
            worksheet[i][8].value = "RESOLVING"
            break

    workbook.save(cfilepath)
    workbook.close()


def MarkReadyStatus(gameID):  # THIS INDICATES THAT THE FUNDS HAVE TRANSFERRED TO THE ESCROW
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']
    
    idx = GetChallengeIdx(gameID)
    
    if (idx != -1):
        escrowID = worksheet[idx][5].value
        escrow = GetEscrowContainer(escrowID)
        
        if (escrow != -1):
            success = True

            assets = escrow['assets']
            
            for i in assets:
                # print("THIS: ", i['status'])
                if i['status'] == 'user_signature_requested' or i['status'] == 'expired':
                    success = False
                    break
                
            if success: worksheet[idx][8].value = "YES" 

    workbook.save(cfilepath)
    workbook.close()    


def HandleFinishedGames():
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']
    
    acceptedGames = []
    finishedGames = []

    for i in range(2, worksheet.max_row + 1):
        if isGameFinished(worksheet[i][0].value):
            finishedGames.append(worksheet[i][0].value)

        if (worksheet[i][6].value != "NO"):
            acceptedGames.append(worksheet[i][0].value)


    for game in acceptedGames:
        MarkReadyStatus(game)
    
    for game in finishedGames:
        if (gameEnded(game) == -1):   # This means escrow was not resolvable (due to insufficient funds)
            MarkResolving(game)
        else:
            FindAndRemoveRow(game)
