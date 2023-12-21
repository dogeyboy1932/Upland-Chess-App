from Chess.FIXED_CHESS_VARIABLES import cfilepath
from openpyxl import load_workbook


def GetChallengeIdx(gameID):
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            return i

    return -1
