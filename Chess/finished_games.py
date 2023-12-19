from openpyxl import load_workbook
from FIXED_CHESS_VARIABLES import cfilepath
from isGameFinished import isGameFinished
from __run_chess_app import gameEnded


def FindAndRemoveRow(gameID):
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            worksheet.delete_rows(i, i)

    workbook.save(cfilepath)
    workbook.close()


def RemoveAndResolveFinishedGames():
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    finishedGames = []

    for i in range(1, worksheet.max_row + 1):
        if isGameFinished(worksheet[i][0].value):
            finishedGames.append(worksheet[i][0].value)

    if not finishedGames:
        return -1

    for game in finishedGames:
        gameEnded(game)
        FindAndRemoveRow(game)

    return finishedGames


# def run():
#     df = pd.read_excel(filepath)
#     print(df)
#
#     RemoveFinishedGames()
#
#
# run()


# def RemovedFinishedGames(finishedGames):
# print("GameID", gameID)
# print("comparable", worksheet[i][0].value)

# print(i)
# print(worksheet[i][0].value)

# print(game)

# import pandas as pd
# df = pd.read_excel(filepath)
# print(df)