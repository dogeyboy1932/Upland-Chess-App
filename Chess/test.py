from FIXED_CHESS_VARIABLES import cfilepath
from FIXED_CHESS_VARIABLES import client
import pandas as pd



# def isGameFinished(gameID):
#     print("THERE")
#     print(gameID)
#     try:
#         client.games.export(gameID)
#         return True
#     except:
#         return False
#     # print("THERE2")

# # AKHIL NOTE: THIS IS PROLLY INEFFICIENT FOR DETERMINING IF A GAME IS OVER OR NOT...MIGHT NEED TO UPDATE

# def run():
#     gameID = "JR0l3LfF"
#     print(isGameFinished(gameID))

# run()

# df = pd.read_excel(cfilepath)
# print(df)

# import berserk

# # Replace 'your_client_id' and 'your_client_secret' with your actual Lichess API credentials
# client_id = 'your_client_id'
# client_secret = 'your_client_secret'

# # Specify the required scopes, e.g., 'board:write'
# scopes = ['board:write']

# # Obtain a token with the specified scopes
# # token = berserk.auth.create_session(client_id, client_secret, scopes)
# token = berserk.Client()

# # from FIXED_CHESS_VARIABLES import token
# # import berserk
# # client = berserk.Client()
# client = berserk.Client(session=token)

# # headers = {"Authorization": f"Bearer {token}"}

# client.board.post_message("1yizMtMm", "TESTING", True)

# # for challenge in challenges:
# #     print(challenge)



from openpyxl import load_workbook
# import pandas as pd
# from FIXED_CHESS_VARIABLES import cfilepath
# from FIXED_CHESS_VARIABLES import client
# # from game_ended import gameEnded



# def isGameFinished(gameID):
#     print("THERE")
#     print(gameID)
#     try:
#         client.games.export(gameID)
#         return True
#     except:
#         return False
    

# def FindAndRemoveRow(gameID):
#     workbook = load_workbook(cfilepath)
#     worksheet = workbook['Sheet']

#     for i in range(1, worksheet.max_row + 1):
#         if worksheet[i][0].value == gameID:
#             worksheet.delete_rows(i)
#             break

#     workbook.save(cfilepath)
#     workbook.close()


# def HandleFinishedGames():
#     workbook = load_workbook(cfilepath)
#     worksheet = workbook['Sheet']

#     finishedGames = []
#     print("HERE1")

#     for i in range(2, worksheet.max_row + 1):
#         if isGameFinished(worksheet[i][0].value):
#             print("here")
#             finishedGames.append(worksheet[i][0].value)

#     if not finishedGames:
#         return -1

#     for game in finishedGames:
#         print(game)
#         # gameEnded(game)
#         FindAndRemoveRow(game)

#     return finishedGames


# def run():
# df = pd.read_excel(cfilepath)
# print(df)

#     HandleFinishedGames()


# def AppendInitial():
#     workbook = load_workbook(cfilepath)
#     worksheet = workbook['Sheet']

#     data = ["gameID", "challenger", "rating", "wager", "link", "escrowID", "accepted?", "accepter"]

#     worksheet.append(data)

#     workbook.save(cfilepath)
#     workbook.close()

# AppendInitial()

# df = pd.read_excel(cfilepath)
# print(df)