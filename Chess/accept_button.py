from openpyxl import load_workbook

from FIXED_VARIABLES import cfilepath
from Chess.query_for_uplandID import QueryForUplandID

from Upland.join_escrow_container import JoinEscrow
from Upland.get_bearer_token import GetBearerToken
from Upland.get_user_profile import GetUserProfile


def ChallengeAccepted(link, challenger, accepter):  # <- Accept Button clicked
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    if GetUserProfile(GetBearerToken(accepter))['level'] == "Visitor":
        return -1
    
    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == link:
            worksheet[i][6].value = True
            worksheet[i][7].value = accepter

            challengeIdx = i

    workbook.save(cfilepath)
    workbook.close()
    
    bearer = GetBearerToken(accepter)
    eid = worksheet[challengeIdx][5].value
    wager = worksheet[challengeIdx][3].value
    
    JoinEscrow(bearer, eid, wager)


    return 1

    # print(bearer, " ", eid, " ", wager)