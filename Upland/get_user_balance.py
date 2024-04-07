from Upland.FIXED_VARIABLES import conn
from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook
import json


def GetUserBalance(upland_access_token):
    payload = ""
    headers = {
        'Authorization': f'Bearer {upland_access_token}',
    }

    conn.request("GET", "/developers-api/user/balances", payload, headers)
    res = conn.getresponse()

    # print(res.read())

    data = json.loads(res.read().decode("utf-8"))

    if res.status == 200:
        return data['availableUpx']
    else:
        print(f'Request failed with status code {res.status}')


def UpdateBalance(upland_access_token):
    payload = ''
    headers = {
        'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiODVmZTQwMTUtYTYwZi00MWQ2LTk0ZmMtMmFlOGIzNWUyMWJlIiwiaWF0IjoxNzEyMzQyOTI5fQ.ka92uTcZJ4XMFYwVNbOqFMcHG3ppPZ7G6IHH9iaqR7U',
        'Cookie': 'sticky-session-1=1712345152.12.32.298740|9a5cc3e4d08faea009d8e16f5c97bee9'
    }
    
    conn.request("GET", "/developers-api/user/balances", payload, headers)
    
    res = conn.getresponse()
    data = json.loads(res.read().decode("utf-8"))

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == upland_access_token:
            worksheet[i][3].value = data['availableUpx']

    workbook.save(filepath)
    workbook.close()


    
def GetUserBalanceOnSheet(uplandID):
    if uplandID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return worksheet[i][3].value

    return -1

# def run():
#     upland_access_token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiNDA5NmJhYTQtNWYzZC00NTY0LTk0M2MtM2IyZTY5YTA1OTg3IiwiaWF0IjoxNzA1NjEzNzcyfQ.lvDEu6WByip5Z8LiFAoWaIGj5MzV8U1uqyRcrgxiuHY"

#     print(UpdateBalance(upland_access_token))

#     # print(GetUserBalanceOnSheet("dogeyboy19"))


# run()


# OLD CODE:
# import requests
# url = 'https://api.sandbox.upland.me/developers-api/user/balances'
#
# headers = {
#     'Authorization': f'Bearer {upland_access_token}',
#     'UserId': upland_user_id,
# }
#
# response = requests.get(url, headers=headers)
#
# if response.status_code == 200:
#     data = response.json()
#     return data['availableUpx']
#
# else:
#     print(f'Request failed with status code {response.status_code}')
