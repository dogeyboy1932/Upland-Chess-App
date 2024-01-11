from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook
from Chess.game_winner import GameWinner

def QueryForEOSID(lichessID):
    if lichessID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == lichessID:
            return worksheet[i][5].value
        
    return -1


def QueryForBearer(uplandID):
    if uplandID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return worksheet[i][4].value

    return -1

def QueryForLichessID(uplandID):
    if uplandID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return worksheet[i][0].value

    return -1


def GetPassword(username):
    if username == "{}":
        return "-1"

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == username:
            return worksheet[i][6].value
        
    return "-1"

def QueryUplandIDRow(uplandID):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return i

    return -1


def run():
    gameID = "rL3ou9hv"
    winner = GameWinner(gameID=gameID)
    corr_uplandID = QueryForEOSID(winner[0])


    print(corr_uplandID)


# run()
