from openpyxl import Workbook
import openpyxl
import pandas as pd

from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook


# def QueryUplandIDRow(uplandID):
#     workbook = load_workbook(filepath)
#     worksheet = workbook['Sheet']
#
#     for i in range(1, worksheet.max_row + 1):
#         if worksheet[i][1].value == uplandID:
#             return worksheet[i][4].value
#
#     return -1
#
#
# print(QueryUplandIDRow("testact1112"))


def run():
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    # Grabbing Details
    print(worksheet[4][0].value)
    print(worksheet[4][1].value)
    print(worksheet[4][2].value)
    print(worksheet[4][3].value)
    print(worksheet[4][4].value)
    print(worksheet[4][5].value)

run()