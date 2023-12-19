from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook
from Chess.chess_game_winner import GameWinner


def QueryForUplandID(lichessID):
    if lichessID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == lichessID:
            return worksheet[i][1].value

    return -1


# def run():
#     gameID = "RLyFq9MX"
#     winner = game_winner(gameID=gameID)
#     corr_uplandID = QueryWinner(winner)
#
#     print(corr_uplandID)
#
#
# run()




# from append_profile import GetWorkbook
# from query_uplandID import QueryUplandID


# print(winner)
# uid = ""
#
# if index != -1:
#     uid = worksheet[index][3].value
#
# print(uid)
# print(QueryUplandID(uid))

# workbook.save(filepath)
# workbook.close()
#
# worksheet = workbook.active
#
# all_rows = list(worksheet.rows)
#
# print(all_rows)
#
# for row in all_rows:
#     print(row)

#        print("COMP ", worksheet[i][1].value)

# workbook = load_workbook(filepath)
# worksheet = workbook['Sheet']
# print(worksheet[i][3].value)
