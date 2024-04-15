from openpyxl import load_workbook
from FIXED_VARIABLES import filepath
from Upland.query_spreadsheet import QueryUplandIDRow


def ReplaceProfileSmall(uplandIdx, lichessID, newBearer):
   workbook = load_workbook(filepath)
   worksheet = workbook['Sheet']

   worksheet[uplandIdx][0].value = lichessID
   worksheet[uplandIdx][4].value = newBearer

   workbook.save(filepath)
   workbook.close()


def ReplaceProfileBig(uplandIdx, lichessID, lichessRating, password):
   workbook = load_workbook(filepath)
   worksheet = workbook['Sheet']

   worksheet[uplandIdx][0].value = lichessID 
   worksheet[uplandIdx][2].value = lichessRating   
   worksheet[uplandIdx][6].value = password 

   workbook.save(filepath)
   workbook.close()


def AppendProfile(data):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    worksheet.append(data)

    workbook.save(filepath)
    workbook.close()