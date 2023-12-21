from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook


def QueryUplandIDRow(uplandID):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return i

    return -1
