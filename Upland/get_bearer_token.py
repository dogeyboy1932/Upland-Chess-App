from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook
import pandas as pd

def GetBearerToken(uplandID):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID.lower():
            return worksheet[i][4].value

    return -1


def run():
    print(GetBearerToken("dogeyboy19"))

# run()
