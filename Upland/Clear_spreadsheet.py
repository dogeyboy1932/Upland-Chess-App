from openpyxl import load_workbook
import pandas as pd
from FIXED_VARIABLES import filepath


def run():
    workbook = load_workbook(filepath)
    sheet = workbook['Sheet']

    print("Maximum rows before removing:", sheet.max_row)

    if sheet.max_row != 0:
        sheet.delete_rows(0, sheet.max_row)

    print("Maximum rows after removing:", sheet.max_row)

    workbook.save(filepath)
    workbook.close()


run()


# save the file to the path
# path = './openpy.xlsx'
# df = pd.read_excel(r"/Users/gogin/Desktop/Metaverse/ChessApp Pycharm Code/ChessDatabase1.xlsx")

# print(df)
# filepath = r"/Users/gogin/Desktop/Metaverse/ChessApp Pycharm Code/ChessDatabase1.xlsx"
# def remove(sheet, row):
#     # iterate the row object
#     sheet.delete_rows(row[0].row)
# select the sheet
# iterate the sheet object
# for row in sheet:
#     remove(sheet, row)