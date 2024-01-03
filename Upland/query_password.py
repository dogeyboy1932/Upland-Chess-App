from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook

def GetPassword(username):
    if username == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == username:
            return worksheet[i][6].value
        
    return -1