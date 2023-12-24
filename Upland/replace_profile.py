from openpyxl import load_workbook
from Upland.FIXED_VARIABLES import filepath
from Upland.query_uplandID_index import QueryUplandIDRow

def ReplaceProfile(userId, profile):
    print("REPLACE CALLED")
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    index = QueryUplandIDRow(userId)

    for col_num, new_value in enumerate(profile, start=1):
       worksheet.cell(row=index, column=col_num, value=new_value)

    workbook.save(filepath)
    workbook.close()