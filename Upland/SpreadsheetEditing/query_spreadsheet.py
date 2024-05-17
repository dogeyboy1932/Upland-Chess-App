from openpyxl import load_workbook
from FIXED_VARIABLES import cfilepath, filepath


def QueryUplandIDRow(uplandID):
    worksheet = load_workbook(filepath)['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return i

    return -1


def QueryForLichessID(uplandID):
    worksheet = load_workbook(filepath)['Sheet']
    idx = QueryUplandIDRow(uplandID)

    return worksheet[idx][0].value if idx != -1 else -1


def GetUserBalanceOnSheet(uplandID):
    worksheet = load_workbook(filepath)['Sheet']
    
    idx = QueryUplandIDRow(uplandID)

    return worksheet[idx][3].value if idx != -1 else -1


def GetBearerToken(uplandID):
    worksheet = load_workbook(filepath)['Sheet']
    
    idx = QueryUplandIDRow(uplandID)

    return worksheet[idx][4].value if idx != -1 else -1


def QueryForPassword(uplandID):
    worksheet = load_workbook(filepath)['Sheet']
    
    idx = QueryUplandIDRow(uplandID)

    return worksheet[idx][6].value if idx != -1 else -1


def QueryForEOSID(lichessID):
    worksheet = load_workbook(filepath)['Sheet']
    
    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == lichessID:
            return worksheet[i][5].value
        
    return -1


def QueryForUplandID(lichessID):
    worksheet = load_workbook(filepath)['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == lichessID:
            return worksheet[i][1].value

    return -1

  

def GetCredentialsByID(userId):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][7].value == userId:
            return [worksheet[i][1].value, worksheet[i][6].value]
        
    return -1


def GetChallengeIdx(gameID):
    workbook = load_workbook(cfilepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][0].value == gameID:
            return i

    return -1

def QueryForIdxByLink(link):
    worksheet = load_workbook(cfilepath)['Sheet']
    
    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == link:
            return i