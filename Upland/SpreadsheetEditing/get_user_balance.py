import json
from openpyxl import load_workbook

from FIXED_VARIABLES import conn, filepath


def GetUserBalance(upland_access_token):
    payload = ""
    headers = {
        'Authorization': f'Bearer {upland_access_token}',
    }

    conn.request("GET", "/developers-api/user/balances", payload, headers)
    res = conn.getresponse()

    # print(res.read())

    data = json.loads(res.read().decode("utf-8"))

    if res.status == 200:
        return data['availableUpx']
    else:
        print(f'Request failed with status code {res.status}')


def UpdateBalance(upland_access_token):
    payload = ''
    headers = {
        'Authorization': f'Bearer {upland_access_token}',
        'Cookie': 'sticky-session-1=1712345152.12.32.298740|9a5cc3e4d08faea009d8e16f5c97bee9'
    }
    
    try:
        conn.request("GET", "/developers-api/user/balances", payload, headers)
        res = conn.getresponse()
    
        data = json.loads(res.read().decode("utf-8"))
        # print(data)
        try: 
            var = data['statusCode']
            return
        except:
            var = 1
    except:
        print("getBalanced call failed")

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][4].value == upland_access_token:
            try:
                worksheet[i][3].value = data['availableUpx']
            except:
                print("availableUpx call failed")

    workbook.save(filepath)
    workbook.close()

    
def GetUserBalanceOnSheet(uplandID):
    if uplandID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return worksheet[i][3].value

    return -1

# def run():
#     upland_access_token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOiI0ODI5OGVhMC0yNDBhLTExZWUtOWMwNC1iMzcyMDk2MTViOGIiLCJhcHBJZCI6MjMyLCJ0b2tlbklkIjoiNDA5NmJhYTQtNWYzZC00NTY0LTk0M2MtM2IyZTY5YTA1OTg3IiwiaWF0IjoxNzA1NjEzNzcyfQ.lvDEu6WByip5Z8LiFAoWaIGj5MzV8U1uqyRcrgxiuHY"

#     print(UpdateBalance(upland_access_token))

#     # print(GetUserBalanceOnSheet("dogeyboy19"))


# run()


# OLD CODE:
# import requests
# url = 'https://api.sandbox.upland.me/developers-api/user/balances'
#
# headers = {
#     'Authorization': f'Bearer {upland_access_token}',
#     'UserId': upland_user_id,
# }
#
# response = requests.get(url, headers=headers)
#
# if response.status_code == 200:
#     data = response.json()
#     return data['availableUpx']
#
# else:
#     print(f'Request failed with status code {response.status_code}')
