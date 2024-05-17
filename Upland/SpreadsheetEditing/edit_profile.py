from openpyxl import load_workbook

from FIXED_VARIABLES import filepath

from Upland.SpreadsheetEditing.query_spreadsheet import QueryUplandIDRow


def FillingLichessInfo(uplandIdx, lichessID, lichessRating, password):
   workbook = load_workbook(filepath)
   worksheet = workbook['Sheet']

   if password == -1: 
      replaceCount = 1 if not worksheet[uplandIdx][8].value else worksheet[uplandIdx][8].value + 1
      worksheet[uplandIdx][8].value = replaceCount   # This count cannot exceed too many times or account is banned
   else:
      worksheet[uplandIdx][6].value = password

   worksheet[uplandIdx][0].value = lichessID 
   worksheet[uplandIdx][2].value = lichessRating   

   workbook.save(filepath)
   workbook.close()


def AppendProfile(data):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    worksheet.append(data)

    workbook.save(filepath)
    workbook.close()


def DeleteProfile(uplandID, password):
    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    idx = QueryUplandIDRow(uplandID)

    if worksheet[idx][6].value != password: return "Incorrect Password"
    worksheet.delete_rows(idx)

    workbook.save(filepath)
    workbook.close()

    return "success"


def AppendProfileHeader():
   data = ["Lichess ID", "Upland Username", "Lichess Rating", "Balance", "Bearer Token", "Eos Upland ID", "Password", "UserId"]
    
   AppendProfile(data)