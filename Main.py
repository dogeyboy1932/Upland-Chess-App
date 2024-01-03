# Hit the run button ^^^

from flask import Flask, request, jsonify
import pandas as pd
import json

from openpyxl import load_workbook

from Upland.FIXED_VARIABLES import credential
from Upland.FIXED_VARIABLES import filepath
from Upland.append_profile import AppendProfile
from Upland.create_profile import CreateProfile
from Upland.query_uplandID_index import QueryUplandIDRow
from Upland.Auth_code import Verify
from Upland.query_password import GetPassword

from Chess.FIXED_CHESS_VARIABLES import NumpyArrayEncoder
from Chess.render_database import Iterate
from Chess.handle_finished_games import HandleFinishedGames
from Chess.challenge_button import ChallengeButtonClicked
from Chess.accept_button import ChallengeAccepted
from Chess.cancel_button import ChallengeCanceled
from Chess.handle_finished_games import ChallengeDeleted
from Chess.handle_finished_games import GetStatusWithLink


from flask_cors import CORS 

app = Flask(__name__)
CORS(app)
app.logger.disabled = True

went = False

@app.route('/database', methods=['POST'])
def ChallengeDatabase():
    # print("DATABASE")
    global went

    if not went:
        arr = Iterate()
        went = True
        print(went)
    else:
        HandleFinishedGames() 
        arr = Iterate()

    encodedNumpyData = json.dumps({"array": arr}, cls=NumpyArrayEncoder)

    return encodedNumpyData


@app.route('/accepted', methods=['POST'])
def Accepted():
    print("Accepted")
    Iterate()

    accepter = request.get_json().get('currentUserUplandID')
    link = request.get_json().get('link')
    challenger = "dogeyboy19"
    
    # print(accepter)
    # print(challenger)

    ChallengeAccepted(link, challenger, accepter)

    return ChallengeDatabase()

    # return jsonify({'message': 'Details received successfully'})


@app.route('/cancel', methods=['POST'])
def Cancel():
    print("Cancel")

    link = request.get_json().get('link')
    
    ChallengeCanceled(link)

    return jsonify({'message': 'Details received successfully'})


@app.route('/delete', methods=['POST'])
def Deleted():
    print("Cancel")

    link = request.get_json().get('link')
    
    # print(str(GetStatusWithLink(link)))

    if (str(GetStatusWithLink(link)) == "False"):
        ChallengeDeleted(link)
        return "Success"
    else:
        return "Already Accepted"

    # return jsonify({'message': 'Details received successfully'})


@app.route('/auth', methods=['POST'])
def Auth():
    res = Verify(credential)

    return res

@app.route('/password', methods=['POST'])
def Password():
    username = request.get_json().get('username')

    realPassword = GetPassword(username)

    print(realPassword)
    return realPassword

@app.route('/credentials', methods=['POST'])
def Credentials():
    username = request.get_json().get('username')
    password = request.get_json().get('password')


    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    prof_pass = -1

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == username:
            prof_pass = worksheet[i][6].value

    if (prof_pass == -1):
        print("HERE")
        return 'no profile found'
    elif (prof_pass == "null"):
        worksheet[i][6].value = password 
    else:
        print("THERE")
        return 'profile exists'        


    workbook.save(filepath)
    workbook.close()

    return 'success'

@app.route('/submit-details', methods=['POST'])
def ChallengeButton():
    print("SUBMITTED")
    data = request.get_json()

    uplandID = data.get('upland')
    rated = data.get('rated')
    wager = data.get('wager')

    print(f"Received details: Name - {rated}, Email - {wager}, UplandID - {uplandID}")

    res = ChallengeButtonClicked(uplandID, rated, wager)

    return jsonify(res)


@app.route('/', methods=['POST'])
def respond():

    data = request.json

    if data['type'] == 'AuthenticationSuccess':
        user_id = data['data']['userId']
        access_token = data['data']['accessToken']

        print(access_token)

        CreateProfile(access_token, user_id)

        df = pd.read_excel(filepath)
        print(df)

    return "success"


def AddInitial():
    data = ["Lichess ID", "Upland Username", "Lichess Rating", "Balance", "Bearer Token", "Eos Upland ID", "Password"]
    if QueryUplandIDRow("Upland Username") == -1:
        AppendProfile(data)


if __name__ == '__main__':
    AddInitial()
    app.run(host="0.0.0.0", port=5000, debug=True)








# print(user_id)
# print(access_token)

# print("MAIN PRINT")
# print("IT WORKS")


# import http.client
#
# conn = http.client.HTTPSConnection('eot5eeu5bgtksf7.m.pipedream.net')
# conn.request("POST", "/", '{"test": "event"}', {'Content-Type': 'application/json'})

# filepath = r"/Users/gogin/Desktop/Metaverse/ChessApp Pycharm Code/ChessDatabase1.xlsx"
