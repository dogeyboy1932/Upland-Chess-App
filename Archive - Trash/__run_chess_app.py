# import Upland.join_escrow_container

# from Chess.FIXED_CHESS_VARIABLES import client

from Chess.FIXED_CHESS_VARIABLES import cfilepath
from Upland.FIXED_VARIABLES import filepath

from Chess.create_open_challenge import CreateOpenChallenge
from Chess.append_challenge import AppendChallenge
from Chess.append_challenge import AppendInitial
# from chess_game_winner import GameWinner
from openpyxl import load_workbook
from Upland.join_escrow_container import JoinEscrow
from Upland.get_bearer_token import GetBearerToken
from Chess.query_for_uplandID import QueryForUplandID
# from Upland.query_uplandID_index import QueryUplandIDRow
from Chess.query_challenge_idx import GetChallengeIdx
# from Upland.query_for_eosId import QueryForEOSID
from Upland.get_escrow_container import GetEscrowContainer
import pandas as pd
from Chess.handle_finished_games import HandleFinishedGames
from Chess.render_database import Iterate
# from Chess.chess_challenge import ChallengeButtonClicked

# cfilepath = "/Users/gogin/Desktop/ChessApp/ChessApp Pycharm Code/ChallengeMap.xlsx"


# FRONTEND DEPENDENT

workbook1 = load_workbook(cfilepath)
chessWorksheet = workbook1['Sheet']

# def ChallengeButtonClicked(uplandID_, rated_, wager_):

#     # print("START")

#     # First check if user is valid...check the uplandDatabase if his profile exists <- This should always be true...
#     # if you create a profile to join the game, your profile should be in there

#     ############################
#     # INSTRUCTION
#     # While loop (while challenge not valid or cancel button isn't clicked pop up shows)
#     # Display popup to extract all parameters

#     challenger = "trashboatsr"  # <- Need to extract this detail on front end
#     speed = "rapid"  # <- Placeholder
#     increment = 0  # <- Placeholder
#     rated = rated_   # <- Placeholder
#     variant = "standard"  # <- Placeholder
#     name = ("Challenge by ", uplandID_)  # <- Placeholder
#     wager = wager_  # <- Placeholder [wager must be less than challenger balance...otherwise]

#     # valid = isChallengeValid(challenger, wager) Queries upland ID and checks balance if wager < balance (it is valid)
#     ############################

#     # print("here")
#     # df = pd.read_excel(cfilepath)
#     # print(df)

#     # Challenge under terms is created
#     thisGame = CreateOpenChallenge(challenger=challenger, speed=speed, increment=increment, variant=variant,
#                                    rated=rated, name=name)

#     # Challenge is appended to spreadsheet (database) + Escrow is created & appended + Challenge ID is extracted
#     gameID = AppendChallenge(challenger, wager, thisGame)

#     # print(thisGame)
#     # print("here")

#     # df = pd.read_excel(cfilepath)
#     # print(df)

#     # print(challenger)


#     # Join Escrow Container of this game
#     uplandID = QueryForUplandID(challenger)
#     bearer = GetBearerToken(uplandID)

#     # print(uplandID)
#     # print(bearer)

#     # If bearer is -1, that means the Upland profile was somehow not added. <- This should never happen

#     challengeIdx = GetChallengeIdx(gameID)
#     # print(challengeIdx)

#     df1 = pd.read_excel(filepath)
#     # print(df1)

#     # print(chessWorksheet[1][0].value)

#     # print("THERE")
#     # print(chessWorksheet.max_row + 1)

#     workbook9 = load_workbook(cfilepath)
#     chessWorksheet1 = workbook9['Sheet']

#     # print(chessWorksheet1[challengeIdx][3].value)

#     eid = str(chessWorksheet1[challengeIdx][5].value)
#     # eid = int(eid)

#     # print("BREAK")

#     # print(eid)
#     # print(wager)
#     # print(bearer)

#     wager = int(wager)
#     # print(eid)

#     JoinEscrow(bearer, eid, wager)

#     print(GetEscrowContainer(eid))

    

    # POP-UP THAT DISPLAYS (GO TO UPLAND ACCOUNT TO ACCEPT WAGER TRANSACTION)
    # User must hit reset button for updates

    ############################
    # INSTRUCTION
    # On front end, there is a database...(render challengemap spreadsheet)
    # Add challenger name + challenger rating + link in game variable
    ############################



# def AcceptButton

    ############################
    # INSTRUCTION
    # AFTER BOTH PLAYERS ACCEPT GAME WILL START...WHEN GAME STARTS DO THIS NEXT
    # Game needs 2 players to start it
    # Need function to let us know when the game starts!!
    # Get lichessId of the challenge creator (Player 1) [Query spreadsheet w/ gameID]
    # challenger cannot be the same lichessID as Player 1 -> otherwise throw a rejection pop-up when clicked
    # CHALLENGER MUST HAVE A VALID LICHESS ID
    # Add challenger lichessID to the challenge [params gameID & challenger]
    # Need to run stream_board_game_state once it does start to get the details of players
    # Call QueryForUplandID and extract respective uplandID of both lichess accounts
    #
    #```
    # Challenger will also be made to join the escrow account when he clicks accept button
    #
    # uplandID = QueryForUplandID(challenger)
    # bearer = GetBearerToken(uplandID)
    #
    # challengeIdx = GetChallengeIdx(gameID)
    # eid = chessWorksheet[challengeIdx][5]
    #
    # wager = chessWorksheet[challengeIdx][3]
    #
    # joinEscrow(bearer, eid, wager)
    # ```
    #
    # POP-UP THAT DISPLAYS (GO TO UPLAND ACCOUNT TO ACCEPT WAGER TRANSACTION)
    # TAKE ME TO CHESS GAME BUTTON -> When clicked get redirect to chessgame link (chessWorksheet[challengeIdx][4])
    # User must hit reset button for updates
    # balances are deducted from accounts

    # Need function to let us know when the game ends!! This triggers gameEnded function
    ############################

# TEST STATEMENT:
def minirun():
    
    added = True

    if not added:
        AppendInitial()


    

    # ChallengeButtonClicked()

# minirun()

AppendInitial()





def resetButtonClicked():
    HandleFinishedGames()
    Iterate()


def run():
    return 0

    ############################
    # INSTRUCTION
    # thisGame = challengeButtonClicked()
    # gameID = thisGame['challenge']['id']

    # If both players join the game on the database, game starts and challenge is removed from database
    # Need an indication that a game has started...only then can this run
    # Maybe create a spreadsheet for in_progress games
    # escrowID = challengeAccepted(gameID)

    # Once game ends, resolve the wager and currencies <- calls function (pass in escrow_ID)
    # gameEnded(gameID)

    ############################


# run()





# gameID = "x5clbBxl"
# print(isGameFinished(gameID=gameID))
#
# gameID = "RLyFq9MX"
# print(isGameFinished(gameID=gameID))
#
# gameID = "QNqY9e75"
# print(isGameFinished(gameID=gameID))

# create_open_ended_challenge()

# gameID = "RLyFq9MX"

# export_game(gameID)

# winner = game_winner(gameID)

# print("Winner: ", winner)

# akhil = Profile(winner, 1000, "dogeyboy19")

# print(akhil.lichessID)
# print(winner)

# if akhil.lichessID == winner:
#     akhil.balanceAdd(100)
#
# print(akhil.balance)


# print(gameID)

# if (client.challenges.accept(gameID)):


# def create_open_ended_challenge(rated, timeLimit, clockIncrement):
#     url = 'https://lichess.org/api/challenge/open'
#
#     token = "lip_ZnYbIrvWC5KDwB8P0wJu"
#
#     headers = {
#         "Authorization": f"Bearer {token}"
#     }
#
#     params = {
#         'rated': str(rated),
#         'clock.limit': str(timeLimit),
#         'clock.increment': str(clockIncrement),
#         'variant': "standard",
#         'name': "Akhil Challenge",
#         'rules': "noRematch",
#     }
#
#     response = requests.post(url=url, headers=headers, params=params)
#     data = response.json()
#
#     if response.status_code == 200:
#         # print(data)
#         return data
#         # return data['challenge']['id']
#     else:
#         print(f'Request failed with status code {response.status_code}')
#
