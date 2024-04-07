from Upland.FIXED_VARIABLES import filepath
from openpyxl import load_workbook

def QueryForLichessID(uplandID):
    if uplandID == "{}":
        return -1

    workbook = load_workbook(filepath)
    worksheet = workbook['Sheet']

    for i in range(1, worksheet.max_row + 1):
        if worksheet[i][1].value == uplandID:
            return worksheet[i][0].value

    return -1


def run():
    lichessID = QueryForLichessID("dogeyboy19")

    print(lichessID)

# run()